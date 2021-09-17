import openpyxl

drive = "resource/Drive.xlsx"
base = "resource/Base.xlsx"
cedulas = []
base_excel_ = {}
base_excel = base_excel_
datos_ = {}


# Obtener objeto de libro
class Excel:

    def __init__(self):
        self.F = []

    def leer_archivo(self, hoja, inicio, fin, file, value=False):
        book = openpyxl.load_workbook(filename=file, data_only=True)
        hoja = book.worksheets[hoja]
        datos_drive = hoja[inicio:fin]
        all_dates = {}

        if value:
            contador = 2
            for fila_ in datos_drive:
                info = [celda.value for celda in fila_]
                all_dates[contador] = info[0:]
                contador += 1

        else:
            for fila_ in datos_drive:
                info = [celda.value for celda in fila_]
                all_dates[info[0]] = info[1:]

        return all_dates

    def add_excel(self, datos, columnas):
        wb = openpyxl.load_workbook(filename=drive)
        ws = wb["BAILE JOVENES"]
        for fila in self.F:
            for column, dato in zip(columnas, datos):
                if type(dato) == list and len(dato) == 3:
                    dato = dato[1]
                elif type(dato) == list and len(dato) != 3:
                    None
                ws.cell(row=fila, column=column, value=dato)
        wb.save(drive)
        print("HORARIOS SELECCIONADOS GUARDADOS EN EXCEL")
        wb.close()
        self.F.clear()

    def buscar_datos(self, disponibilidad, x):
        for i in disponibilidad:
            if x == disponibilidad[i][0]:
                self.F.append(i)
                break

    def save_excel(self, historial, disponibilidad, valores):
        for i in list(historial.values())[0][7]:
            self.buscar_datos(disponibilidad, i)
        self.add_excel(list(historial.values())[0], valores)


#  Objeto del archivo Base.xlsx
EXCEL = Excel()
base_excel = EXCEL.leer_archivo(1, "A2", "F82", base) #Cambié la G por la F cualquier error venir y corregir
prueba = EXCEL.leer_archivo(3, "A2", "F2", base) #Cambié la G por la F cualquier error venir y corregir

cedulas_ = list(base_excel.keys())

def horas_pendientes():
    base_p = EXCEL.leer_archivo(1, "A2", "E82", base)
    for ip in base_p:
        base_excel_[str(ip)] = base_p[ip]
    return base_excel_


for i in cedulas_:
    cedulas.append(str(i))


def disponibilidad():
    disponibilidad_ = EXCEL.leer_archivo(6, "A2", "V256", drive, True)
    return disponibilidad_


excel_ba = EXCEL.leer_archivo(1, "C2", "C82", base)

def phone_to_id_convertion(historial):
    id_to_cero = {}
    for i in historial:
        id_to_cero[historial[i][1]] = 0
    return id_to_cero

def add_cero_to_base(H_):
    wb = openpyxl.load_workbook(filename=base)
    ws = wb["Hoja 2"]
    dato = 0
    column = 5
    for p, z in zip(excel_ba, range(2, len(excel_ba))):
        if p == list(H_.values())[0][1]:
            fila = z
            ws.cell(row=fila, column=column, value=dato)
            wb.save(base)
            wb.close()
            break

def fecha(dato):
    resultado = ""
    meses = {"01": "enero", "02": "febrero", "03": "marzo", "04":"abril", "05": "mayo", "06": "junio", "07": "julio",
             "08": "agosto", "09": "septiembre", "10": "octubre", "11": "noviembre", "10": "diciembre"}
    mes = ""
    for mes_ in meses:
        if mes_ == dato[5:7]:
            mes = meses[mes_]
    resultado = dato[8:10] + " de " + mes + " del " + dato[0:4]
    return resultado

def extraer_mes(dato):
    meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
             "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    for i in meses:
        if i in dato:
            return i



def consulta():
    consulta_ = EXCEL.leer_archivo(6, "A2", "Q184", drive)
    datos_consulta = {}

    for i in consulta_:
        if i != "ok":
            #datos_consulta[i] = [consulta_[i][9] + " " + fecha(str(consulta_[i][14])[0:10]), str(consulta_[i][15])[0:5]] #, consulta_[i][8]
            datos_consulta[i] = [str(consulta_[i][14])[0:-7]+"a las " + str(consulta_[i][15])[0:5] + " a.m."]
    return datos_consulta

def datos_correo(data):
    consulta_ = EXCEL.leer_archivo(6, "A2", "Q184", drive, value=True)
    datos = {}
    for i in consulta_:
        if consulta_[i][3] == list(data.values())[0][1]:
            datos[i] = ["Avenida 68", consulta_[i][12], consulta_[i][13], consulta_[i][14], str(consulta_[i][15])[0:-7], str(consulta_[i][16])[0:5]]
    return datos

def add_excel_base(phone, valor):
    wb = openpyxl.load_workbook(filename=base)
    ws = wb["Hoja 2"]
    contador = 2
    for i in prueba:
        if str(prueba[i][4]) == phone:
            ws.cell(row=contador, column=9, value=valor)
            wb.save(base)
            wb.close()
            print("GUARDADO EN EXCEL: ", phone, "==> ", valor)
            break
        contador +=1




#add_excel_base("3156567112", "sedejjjj")
